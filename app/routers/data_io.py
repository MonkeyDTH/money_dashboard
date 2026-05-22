from fastapi import APIRouter, Request, Depends
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select
import io

from app.database import get_session
from app.deps import templates
from app.models import Account, AccountSnapshot, Company, SalaryRecord, SalaryItem, Member

router = APIRouter(prefix="/data", tags=["data"])


@router.get("", name="data_io")
async def data_io_page(request: Request):
    return templates.TemplateResponse(request, "data_io/index.html", {})


@router.get("/export", name="data_export")
async def data_export(session: Session = Depends(get_session)):
    from openpyxl import Workbook
    from app.models.salary import ITEM_TYPE_LABELS

    wb = Workbook()

    ws = wb.active
    ws.title = "accounts"
    ws.append(["id", "名称", "类型", "归属", "货币", "是否激活", "备注"])
    for a in session.exec(select(Account)).all():
        ws.append([a.id, a.name, a.type.value, a.owner_type.value, a.currency, a.is_active, a.note])

    ws2 = wb.create_sheet("snapshots")
    ws2.append(["account_id", "账户名", "月份(period)", "余额", "录入时间"])
    for s in session.exec(select(AccountSnapshot).order_by(AccountSnapshot.period)).all():
        acc = session.get(Account, s.account_id)
        ws2.append([s.account_id, acc.name if acc else "", str(s.period), float(s.balance), str(s.recorded_at)])

    ws3 = wb.create_sheet("companies")
    ws3.append(["id", "公司名", "member_id", "入职日期", "离职日期", "备注"])
    for c in session.exec(select(Company)).all():
        ws3.append([c.id, c.name, c.member_id, str(c.start_date), str(c.end_date) if c.end_date else "", c.note])

    ws4 = wb.create_sheet("salaries")
    ws4.append(["id", "member_id", "company_id", "月份(period)", "发薪日", "应发", "实发", "备注"])
    for r in session.exec(select(SalaryRecord).order_by(SalaryRecord.period)).all():
        ws4.append([r.id, r.member_id, r.company_id, str(r.period), str(r.pay_date) if r.pay_date else "", float(r.gross), float(r.net), r.note])

    ws5 = wb.create_sheet("salary_items")
    ws5.append(["salary_record_id", "明细类型", "自定义名称", "金额", "方向"])
    for item in session.exec(select(SalaryItem)).all():
        ws5.append([
            item.salary_record_id,
            ITEM_TYPE_LABELS.get(item.item_type, item.item_type.value),
            item.label,
            float(item.amount),
            item.direction.value,
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=money_dashboard_export.xlsx"},
    )
